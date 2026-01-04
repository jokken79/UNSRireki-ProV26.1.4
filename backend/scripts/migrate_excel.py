#!/usr/bin/env python3
"""
Migration script for Excel 社員台帳 (Employee Register) to PostgreSQL.

Migrates data from:
- DBStaffX (Master Employee Data)
- DBGenzaiX (派遣社員 - Dispatch Workers)
- DBUkeoiX (請負社員 - Contract Workers)

Usage:
    python migrate_excel.py --excel-file "path/to/社員台帳.xlsm" --output-dir "./exports"
"""
import os
import sys
import json
import argparse
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Any, Optional


def parse_args():
    parser = argparse.ArgumentParser(
        description="Migrate Excel 社員台帳 to PostgreSQL"
    )
    parser.add_argument(
        "--excel-file",
        required=True,
        help="Path to Excel file (.xlsx or .xlsm)"
    )
    parser.add_argument(
        "--output-dir",
        default="./exports",
        help="Directory to export extracted data"
    )
    return parser.parse_args()


def load_excel(file_path: str):
    """Load Excel file using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    try:
        return openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        sys.exit(1)


def parse_date(value) -> Optional[str]:
    """Parse various date formats to ISO string."""
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')

    # Try to parse string dates
    if isinstance(value, str):
        for fmt in ['%Y/%m/%d', '%Y-%m-%d', '%Y年%m月%d日']:
            try:
                return datetime.strptime(value, fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue

    # Excel serial date number
    if isinstance(value, (int, float)) and value > 25000:
        try:
            from datetime import timedelta
            base_date = datetime(1899, 12, 30)
            return (base_date + timedelta(days=int(value))).strftime('%Y-%m-%d')
        except:
            pass

    return None


def extract_staff_master(wb) -> List[Dict]:
    """Extract DBStaffX (Master Employee Data) - 17 columns."""
    sheet_name = 'DBStaffX'
    if sheet_name not in wb.sheetnames:
        print(f"Warning: Sheet '{sheet_name}' not found")
        return []

    ws = wb[sheet_name]
    employees = []

    # Column mapping (1-indexed to match Excel)
    columns = {
        1: 'status',           # №
        2: 'employee_number',  # 社員№
        3: 'office',           # 事務所
        4: 'full_name',        # 氏名
        5: 'name_kana',        # カナ
        6: 'gender',           # 性別
        7: 'nationality',      # 国籍
        8: 'birth_date',       # 生年月日
        9: 'age',              # 年齢
        10: 'visa_expiry',     # ビザ期限
        11: 'visa_type',       # ビザ種類
        12: 'has_spouse',      # 配偶者
        13: 'postal_code',     # 〒
        14: 'address',         # 住所
        15: 'building_name',   # 建物名
        16: 'hire_date',       # 入社日
        17: 'termination_date' # 退社日
    }

    # Skip header row
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[1]:  # Skip empty rows (no employee number)
            continue

        employee = {}
        for col_idx, key in columns.items():
            value = row[col_idx - 1] if col_idx <= len(row) else None

            # Parse dates
            if key in ['birth_date', 'visa_expiry', 'hire_date', 'termination_date']:
                value = parse_date(value)

            # Parse boolean
            if key == 'has_spouse':
                value = value == '有' if isinstance(value, str) else bool(value)

            employee[key] = value

        employees.append(employee)

    return employees


def extract_haken(wb) -> List[Dict]:
    """Extract DBGenzaiX (派遣社員) - 42 columns."""
    sheet_name = 'DBGenzaiX'
    if sheet_name not in wb.sheetnames:
        print(f"Warning: Sheet '{sheet_name}' not found")
        return []

    ws = wb[sheet_name]
    assignments = []

    # Column mapping for key fields
    columns = {
        1: 'status',                # 現在
        2: 'employee_number',       # 社員№
        3: 'client_company_id',     # 派遣先ID
        4: 'client_company',        # 派遣先
        5: 'assignment_location',   # 配属先
        6: 'assignment_line',       # 配属ライン
        7: 'job_description',       # 仕事内容
        8: 'full_name',             # 氏名
        14: 'hourly_rate',          # 時給
        15: 'hourly_rate_history',  # 時給改定
        16: 'billing_rate',         # 請求単価
        17: 'billing_rate_history', # 請求改定
        18: 'profit_margin',        # 差額利益
        19: 'standard_salary',      # 標準報酬
        20: 'health_insurance',     # 健康保険
        21: 'nursing_insurance',    # 介護保険
        22: 'pension',              # 厚生年金
        23: 'visa_expiry',          # ビザ期限
        28: 'apartment_name',       # ｱﾊﾟｰﾄ
        29: 'move_in_date',         # 入居
        30: 'start_date',           # 入社日
        31: 'end_date',             # 退社日
        33: 'social_insurance_enrolled',  # 社保加入
        37: 'license_type',         # 免許種類
        38: 'license_expiry',       # 免許期限
        39: 'commute_method',       # 通勤方法
        41: 'japanese_certification' # 日本語検定
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[1]:  # Skip empty rows
            continue

        assignment = {'employment_type': 'haken'}
        for col_idx, key in columns.items():
            value = row[col_idx - 1] if col_idx <= len(row) else None

            # Parse dates
            if key in ['visa_expiry', 'move_in_date', 'start_date', 'end_date', 'license_expiry']:
                value = parse_date(value)

            # Parse numbers
            if key in ['hourly_rate', 'billing_rate', 'profit_margin', 'standard_salary',
                      'health_insurance', 'nursing_insurance', 'pension']:
                value = int(value) if value and str(value).isdigit() else None

            assignment[key] = value

        assignments.append(assignment)

    return assignments


def extract_ukeoi(wb) -> List[Dict]:
    """Extract DBUkeoiX (請負社員) - 36 columns."""
    sheet_name = 'DBUkeoiX'
    if sheet_name not in wb.sheetnames:
        print(f"Warning: Sheet '{sheet_name}' not found")
        return []

    ws = wb[sheet_name]
    assignments = []

    # Column mapping for key fields
    columns = {
        1: 'status',               # 現在
        2: 'employee_number',      # 社員№
        3: 'job_type',             # 請負業務
        4: 'full_name',            # 氏名
        10: 'hourly_rate',         # 時給
        11: 'hourly_rate_history', # 時給改定
        16: 'commute_distance',    # 通勤距離
        17: 'transport_allowance', # 交通費
        18: 'profit_margin',       # 差額利益
        24: 'apartment_name',      # ｱﾊﾟｰﾄ
        25: 'move_in_date',        # 入居
        26: 'start_date',          # 入社日
        27: 'end_date',            # 退社日
        30: 'bank_account_name',   # 口座名義
        31: 'bank_name',           # 銀行名
        32: 'branch_number',       # 支店番号
        33: 'branch_name',         # 支店名
        34: 'account_number',      # 口座番号
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[1]:  # Skip empty rows
            continue

        assignment = {'employment_type': 'ukeoi'}
        for col_idx, key in columns.items():
            value = row[col_idx - 1] if col_idx <= len(row) else None

            # Parse dates
            if key in ['move_in_date', 'start_date', 'end_date']:
                value = parse_date(value)

            # Parse numbers
            if key in ['hourly_rate', 'transport_allowance', 'profit_margin']:
                value = int(value) if value and str(value).isdigit() else None

            # Parse decimal
            if key == 'commute_distance':
                try:
                    value = float(value) if value else None
                except:
                    value = None

            assignment[key] = value

        assignments.append(assignment)

    return assignments


def save_json(data: list, output_path: str):
    """Save extracted data to JSON file."""
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)

    print(f"Saved {len(data)} records to: {output_path}")


def main():
    args = parse_args()

    print(f"Loading Excel file: {args.excel_file}")
    wb = load_excel(args.excel_file)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Extract and save DBStaffX
    print("\nExtracting DBStaffX (Master Employee Data)...")
    staff_master = extract_staff_master(wb)
    print(f"Found {len(staff_master)} employees")
    save_json(staff_master, str(output_dir / "staff_master.json"))

    # Extract and save DBGenzaiX
    print("\nExtracting DBGenzaiX (派遣社員)...")
    haken = extract_haken(wb)
    print(f"Found {len(haken)} dispatch workers")
    save_json(haken, str(output_dir / "haken_assignments.json"))

    # Extract and save DBUkeoiX
    print("\nExtracting DBUkeoiX (請負社員)...")
    ukeoi = extract_ukeoi(wb)
    print(f"Found {len(ukeoi)} contract workers")
    save_json(ukeoi, str(output_dir / "ukeoi_assignments.json"))

    print("\nMigration completed successfully!")
    print(f"Output directory: {output_dir}")


if __name__ == "__main__":
    main()
